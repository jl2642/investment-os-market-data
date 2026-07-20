from __future__ import annotations

import hashlib
import io
import time
from typing import Any

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

SZSE_XLSX_URL = "https://www.szse.cn/api/report/ShowReport"
SZSE_JSON_URL = "https://www.szse.cn/api/report/ShowReport/data"
SZSE_REFERER = "https://www.szse.cn/szhk/hkbussiness/underlylist/"


def configure_retries(session: requests.Session) -> requests.Session:
    retry = Retry(total=8, connect=8, read=8, status=8, backoff_factor=1.0, status_forcelist=[429, 500, 502, 503, 504], allowed_methods=frozenset(["GET"]))
    adapter = HTTPAdapter(max_retries=retry, pool_connections=2, pool_maxsize=2)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def _sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for ws in wb.worksheets:
        for values in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in values]
            code_idx = next((i for i, v in enumerate(cells) if v.isdigit() and 1 <= len(v) <= 5), None)
            if code_idx is None:
                continue
            code = cells[code_idx].zfill(5)
            if len(code) != 5:
                continue
            cn = cells[code_idx + 1] if code_idx + 1 < len(cells) else ""
            en = cells[code_idx + 2] if code_idx + 2 < len(cells) else ""
            if not cn or cn in {"中文简称", "证券简称"}:
                continue
            rows.append({"stock_code": code, "name_cn": cn, "name_en": en, "security_type": "", "trade_flag": "1"})
    dedup = {x["stock_code"]: x for x in rows}
    return [dedup[x] for x in sorted(dedup)]


def fetch_szse_robust(session: requests.Session) -> tuple[list[dict[str, str]], dict[str, Any], bytes]:
    configure_retries(session)
    params = {"SHOWTYPE": "xlsx", "CATALOGID": "SGT_GGTBDQD", "TABKEY": "tab1", "random": str(time.time())}
    try:
        response = session.get(SZSE_XLSX_URL, params=params, headers={"Referer": SZSE_REFERER}, timeout=90)
        response.raise_for_status()
        rows = _parse_xlsx(response.content)
        if len(rows) >= 500:
            # Obtain only metadata page for the official update date.
            meta_response = session.get(SZSE_JSON_URL, params={"SHOWTYPE":"JSON","CATALOGID":"SGT_GGTBDQD","TABKEY":"tab1","PAGENO":"1","random":str(time.time())}, headers={"Referer":SZSE_REFERER}, timeout=60)
            meta_response.raise_for_status()
            metadata = meta_response.json()[0]["metadata"]
            expected = int(metadata.get("recordcount") or 0)
            if expected and expected != len(rows):
                raise ValueError(f"SZSE_XLSX_RECORD_COUNT:{len(rows)}!={expected}")
            return rows, {"source":"SZSE","url":response.url,"update_date":str(metadata.get("subname") or ""),"record_count":len(rows),"page_count":1,"transport":"OFFICIAL_XLSX","raw_sha256":_sha(response.content)}, response.content
    except Exception:
        pass

    all_rows: list[dict[str, str]] = []
    raw_pages: list[bytes] = []
    first_meta = None
    page = 1
    while True:
        response = session.get(SZSE_JSON_URL, params={"SHOWTYPE":"JSON","CATALOGID":"SGT_GGTBDQD","TABKEY":"tab1","PAGENO":str(page),"random":str(time.time())}, headers={"Referer":SZSE_REFERER,"Connection":"close"}, timeout=90)
        response.raise_for_status()
        raw_pages.append(response.content)
        block = response.json()[0]
        metadata = block.get("metadata") or {}
        if first_meta is None:
            first_meta = metadata
        for x in block.get("data") or []:
            all_rows.append({"stock_code":str(x.get("zqdm","")).zfill(5),"name_cn":str(x.get("zqjc","")).strip(),"name_en":str(x.get("zqywjc","")).strip(),"security_type":"","trade_flag":"1"})
        if page >= int(metadata.get("pagecount") or 0):
            break
        page += 1
        time.sleep(0.35)
    expected = int((first_meta or {}).get("recordcount") or 0)
    if len(all_rows) != expected or expected < 500:
        raise ValueError(f"SZSE_JSON_RECORD_COUNT:{len(all_rows)}!={expected}")
    raw = b"\n".join(raw_pages)
    return all_rows, {"source":"SZSE","url":SZSE_JSON_URL,"update_date":str((first_meta or {}).get("subname") or ""),"record_count":len(all_rows),"page_count":page,"transport":"OFFICIAL_JSON_PAGINATED","raw_sha256":_sha(raw)}, raw
